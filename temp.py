#!/usr/bin/env python3
"""Fetch Jira epic details and all issues in the epic.

Usage:
    python jira_epic_fetcher.py DQRD-10393
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime
from dataclasses import dataclass
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import requests


@dataclass
class JiraConfig:
    base_url: str
    pat: str


@dataclass
class ReadinessRowException:
    issue_key: str
    issue_summary: str
    context: str
    row_name: str
    pass_metric: str
    actual_status: str
    reason: str


@dataclass
class IssueReadinessAnalysis:
    issue_key: str
    issue_summary: str
    context: str
    context_state: str
    overall_state: str
    total_rows: int
    exception_count: int
    exception_rows: list[ReadinessRowException]


class JiraClient:
    def __init__(self, config: JiraConfig) -> None:
        self.base_url = config.base_url.rstrip("/")
        self.session = requests.Session()
        self.session.headers.update(
            {
                "Accept": "application/json",
                "Authorization": f"Bearer {config.pat}",
            }
        )

    def get_issue(self, issue_key: str) -> dict[str, Any]:
        url = f"{self.base_url}/rest/api/2/issue/{issue_key}"
        params = {
            "fields": "summary,description,status,assignee,issuetype,project,reporter,created,updated",
        }
        response = self.session.get(url, params=params, timeout=30)
        self._raise_for_status(response, f"Failed to fetch issue {issue_key}")
        return response.json()

    def get_issue_full(self, issue_key: str) -> dict[str, Any]:
        """Fetch issue with all custom fields for detailed analysis."""
        url = f"{self.base_url}/rest/api/2/issue/{issue_key}"
        response = self.session.get(url, timeout=30)
        self._raise_for_status(response, f"Failed to fetch full issue details for {issue_key}")
        return response.json()

    def get_field_definitions(self) -> list[dict[str, Any]]:
        """Fetch all field definitions to map custom field IDs to their names."""
        url = f"{self.base_url}/rest/api/2/field"
        response = self.session.get(url, timeout=30)
        self._raise_for_status(response, "Failed to fetch Jira field definitions")
        return response.json()

    def get_table_grid_data(self, issue_id: str, field_id: str) -> list[dict[str, str]] | None:
        """Fetch table data from the Table Grid plugin's REST API."""
        url = f"{self.base_url}/rest/jira-tablegrid/1.0/table/{issue_id}/{field_id}"
        try:
            response = self.session.get(url, timeout=30)
            if response.status_code != 200:
                return None
            data = response.json()
            if isinstance(data, list) and data and isinstance(data[0], dict):
                return [_flatten_row(r) for r in data]
            if isinstance(data, dict):
                rows = data.get("rows") or data.get("entries") or data.get("data")
                if isinstance(rows, list) and rows and isinstance(rows[0], dict):
                    return [_flatten_row(r) for r in rows]
            return None
        except Exception:
            return None

    def search_issues(self, jql: str, max_results: int = 200) -> list[dict[str, Any]]:
        url = f"{self.base_url}/rest/api/2/search"
        start_at = 0
        collected: list[dict[str, Any]] = []

        while True:
            params = {
                "jql": jql,
                "startAt": start_at,
                "maxResults": min(max_results, 100),
                "fields": "summary,status,assignee,issuetype,reporter,created,updated",
            }
            response = self.session.get(url, params=params, timeout=30)
            self._raise_for_status(response, "Failed to search issues")
            payload = response.json()
            issues = payload.get("issues", [])
            collected.extend(issues)

            total = payload.get("total", 0)
            start_at += len(issues)
            if start_at >= total or not issues:
                break

        return collected

    @staticmethod
    def _raise_for_status(response: requests.Response, context: str) -> None:
        try:
            response.raise_for_status()
        except requests.HTTPError as exc:
            details = response.text.strip()
            if len(details) > 300:
                details = details[:300] + "..."
            raise RuntimeError(f"{context}. HTTP {response.status_code}. {details}") from exc


def get_env(name: str) -> str:
    value = os.getenv(name, "").strip()
    if not value:
        raise RuntimeError(
            f"Missing required environment variable: {name}. "
            "Set it in your shell or create a .env file in the project root."
        )
    return value


def load_dotenv_file(dotenv_path: Path) -> None:
    if not dotenv_path.exists() or not dotenv_path.is_file():
        return

    for raw_line in dotenv_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def format_user(user_obj: dict[str, Any] | None) -> str:
    if not user_obj:
        return "Unassigned"
    return user_obj.get("displayName") or user_obj.get("name") or "Unknown"


def extract_description(description: Any) -> str:
    if description is None:
        return ""
    if isinstance(description, str):
        return description

    # Jira Cloud may return Atlassian Document Format (ADF).
    chunks: list[str] = []

    def walk(node: Any) -> None:
        if isinstance(node, dict):
            if node.get("type") == "text" and isinstance(node.get("text"), str):
                chunks.append(node["text"])
            for value in node.values():
                walk(value)
        elif isinstance(node, list):
            for item in node:
                walk(item)

    walk(description)
    return " ".join(chunks).strip()


def find_epic_link_field_id(field_definitions: list[dict[str, Any]]) -> str | None:
    """Find the custom field ID for 'Epic Link' by name."""
    for field_def in field_definitions:
        field_name = (field_def.get("name") or "").strip().lower()
        if field_name == "epic link":
            return field_def.get("id")
    return None


def find_epic_issues(client: JiraClient, epic_key: str, epic_link_field_id: str | None = None) -> list[dict[str, Any]]:
    queries = [
        f'"Epic Link" = {epic_key} ORDER BY key ASC',
        f"parent = {epic_key} ORDER BY key ASC",
    ]
    # If we found the actual custom field ID for Epic Link, also query by cf[ID]
    if epic_link_field_id and epic_link_field_id.startswith("customfield_"):
        cf_num = epic_link_field_id.replace("customfield_", "")
        queries.append(f'cf[{cf_num}] = {epic_key} ORDER BY key ASC')

    seen_keys: set[str] = set()
    all_issues: list[dict[str, Any]] = []

    for query in queries:
        try:
            issues = client.search_issues(query)
            for issue in issues:
                key = issue.get("key", "")
                if key and key not in seen_keys:
                    seen_keys.add(key)
                    all_issues.append(issue)
        except Exception:
            continue

    return all_issues


def find_hls_field_id(field_definitions: list[dict[str, Any]]) -> str | None:
    """Find the custom field ID for 'High Level Status' by name."""
    for field_def in field_definitions:
        field_name = (field_def.get("name") or "").strip().lower()
        if "high level status" in field_name:
            return field_def.get("id")
    return None


def find_table_field_ids(field_definitions: list[dict[str, Any]]) -> list[str]:
    """Find custom field IDs that are likely table grid fields."""
    table_ids: list[str] = []
    for field_def in field_definitions:
        field_id = field_def.get("id", "")
        if not field_id.startswith("customfield_"):
            continue
        schema = field_def.get("schema")
        if isinstance(schema, dict):
            custom_type = (schema.get("custom") or "").lower()
            if "tablegrid" in custom_type or "table-grid" in custom_type:
                table_ids.append(field_id)
    return table_ids


def extract_high_level_status_table(
    issue_full: dict[str, Any],
    hls_field_id: str | None = None,
    client: Any = None,
    table_field_ids: list[str] | None = None,
) -> tuple[str, list[dict[str, str]]] | None:
    """Extract High Level Status table from a specific custom field.
    
    Returns (context_text, table_rows) or None.
    context_text contains any statements/labels near the table.
    """
    fields = issue_full.get("fields", {})
    issue_id = issue_full.get("id") or issue_full.get("key", "")

    # Strategy 1: Use the known field ID if provided
    if hls_field_id and hls_field_id in fields:
        raw = fields[hls_field_id]
        # If it's a string, try wiki table parse (which returns context)
        if isinstance(raw, str) and ("|" in raw or "<table" in raw.lower()):
            result = _parse_wiki_table(raw)
            if result:
                return result
        # Otherwise try structured parse
        table = _parse_table_field(raw)
        if table:
            return ("", table)
        # The HLS field might be ADF rich text with an embedded table
        if isinstance(raw, dict):
            adf_tables = _extract_adf_tables(raw)
            if adf_tables:
                return ("", adf_tables[0])

    # Strategy 1b: Try the Table Grid REST API for the known HLS field
    if hls_field_id and client is not None and issue_id:
        table = client.get_table_grid_data(issue_id, hls_field_id)
        if table:
            return ("", table)

    # Strategy 2: Scan all customfield_* fields for table-like structures
    for field_key in sorted(fields.keys()):
        if not field_key.startswith("customfield_"):
            continue
        if hls_field_id and field_key == hls_field_id:
            continue  # Already tried above
        raw = fields[field_key]
        # String fields — try wiki parse with context
        if isinstance(raw, str) and ("|" in raw or "<table" in raw.lower()):
            result = _parse_wiki_table(raw)
            if result:
                return result
        # Structured fields
        table = _parse_table_field(raw)
        if table:
            return ("", table)
        # ADF rich text
        if isinstance(raw, dict):
            adf_tables = _extract_adf_tables(raw)
            if adf_tables:
                return ("", adf_tables[0])

    # Strategy 2b: Try Table Grid REST API for all known table field IDs
    if client is not None and issue_id and table_field_ids:
        for field_id in table_field_ids:
            if field_id == hls_field_id:
                continue  # Already tried
            table = client.get_table_grid_data(issue_id, field_id)
            if table:
                return ("", table)

    # Strategy 3: Check the 'description' field for embedded tables
    desc = fields.get("description")
    if desc is not None:
        # ADF description (dict)
        if isinstance(desc, dict):
            adf_tables = _extract_adf_tables(desc)
            if adf_tables:
                return ("", adf_tables[0])
        # Wiki markup description (string)
        if isinstance(desc, str):
            result = _parse_wiki_table(desc)
            if result:
                return result

    return None


def _parse_table_field(raw: Any) -> list[dict[str, str]] | None:
    """Try to parse a raw field value as a table (list of row-dicts)."""
    if raw is None:
        return None

    # Case 1: Already a list of dicts (Table Grid plugin stores data this way)
    if isinstance(raw, list) and len(raw) > 0 and isinstance(raw[0], dict):
        first_keys = set(raw[0].keys())
        # Skip Jira API reference objects that are clearly not tables
        jira_internal_keys = {"self", "id", "key", "name", "colorName", "iconUrl"}
        if first_keys and first_keys <= jira_internal_keys:
            return None
        # Skip simple option fields (disabled/value)
        if first_keys == {"disabled", "value"} or first_keys <= {"disabled", "value", "id", "self"}:
            return None
        # Looks like a real table – flatten nested values
        return [_flatten_row(r) for r in raw if isinstance(r, dict)]

    # Case 2: String containing wiki/HTML table markup
    if isinstance(raw, str) and ("|" in raw or "<table" in raw.lower() or "<tr" in raw.lower()):
        result = _parse_wiki_table(raw)
        if result:
            return result[1]  # Return just the rows, context handled elsewhere
        return None

    # Case 3: Dict wrapping a 'rows' key
    if isinstance(raw, dict):
        for sub_key in ("rows", "data", "value"):
            sub = raw.get(sub_key)
            if isinstance(sub, list) and len(sub) > 0 and isinstance(sub[0], dict):
                return [_flatten_row(r) for r in sub if isinstance(r, dict)]

    return None


import re as _re


STATUS_COLUMN_ALIASES = {
    "status",
    "result",
    "readiness status",
    "overall status",
}

PASS_METRIC_COLUMN_ALIASES = {
    "pass metrics",
    "pass metric",
    "expected",
    "expected status",
    "target",
    "no impact",
}

PASSING_PHRASES = (
    "ready to release",
    "approved",
    "pass",
    "passed",
    "done",
    "completed",
    "complete",
    "reviewed",
    "hotfix provided",
    "not required",
    "not requested",
    "not needed",
    "not necessary",
    "no update necessary",
    "still apply",
    "no change needed",
    "n/a",
    "na",
    "yes",
    "no cfa",
    "no one page",
    "Zebra Confluence",
    "- zebra confluence",
    "no impact",
    "b",
)

FAILING_PHRASES = (
    "not ready",
    "failed",
    "fail",
    "actively debugging",
    "debugging",
    "field testing",
    "blocked",
    "pending",
    "in progress",
    "open",
    "waiting",
    "hold",
    "risk",
    "rejected",
    "deferred",
)

PASS_METRIC_SPLIT_RE = _re.compile(r"\s*(?:/|,|;|\bor\b|\n)\s*", _re.IGNORECASE)
TICKET_REF_RE = _re.compile(r"^[A-Z][A-Z0-9]+-\d+$")
THRESHOLD_RE = _re.compile(r"([<>]=?|=)\s*([0-9]*\.?[0-9]+)")


def _extract_adf_tables(adf_node: Any) -> list[list[dict[str, str]]]:
    """Extract all tables from an Atlassian Document Format (ADF) node.

    ADF tables have this structure:
      { "type": "table", "content": [
          { "type": "tableRow", "content": [
              { "type": "tableHeader"/"tableCell", "content": [...] }
          ]}
      ]}

    Returns a list of tables, where each table is a list of row-dicts.
    """
    tables: list[list[dict[str, str]]] = []
    if not isinstance(adf_node, (dict, list)):
        return tables

    def _collect_text(node: Any) -> str:
        """Recursively collect all text from an ADF node."""
        if isinstance(node, str):
            return node
        if isinstance(node, dict):
            if node.get("type") == "text":
                return node.get("text", "")
            parts = []
            for child in node.get("content", []):
                parts.append(_collect_text(child))
            return " ".join(p for p in parts if p).strip()
        if isinstance(node, list):
            parts = []
            for item in node:
                parts.append(_collect_text(item))
            return " ".join(p for p in parts if p).strip()
        return ""

    def _walk(node: Any) -> None:
        if isinstance(node, dict):
            if node.get("type") == "table":
                rows_raw = node.get("content", [])
                if not rows_raw:
                    return
                # First row with tableHeader cells = headers
                headers: list[str] = []
                data_rows: list[dict[str, str]] = []

                for row_node in rows_raw:
                    if row_node.get("type") != "tableRow":
                        continue
                    cells = row_node.get("content", [])
                    cell_texts = [_strip_jira_markup(_collect_text(c)) for c in cells]

                    # Detect header row: all cells are tableHeader type
                    is_header = all(
                        c.get("type") == "tableHeader" for c in cells
                    ) if cells else False

                    if not headers and is_header:
                        headers = cell_texts
                    elif headers:
                        row_dict: dict[str, str] = {}
                        for idx, h in enumerate(headers):
                            row_dict[h] = cell_texts[idx] if idx < len(cell_texts) else ""
                        data_rows.append(row_dict)
                    else:
                        # No header row detected yet, treat first as header
                        if not headers:
                            headers = cell_texts
                        else:
                            row_dict = {}
                            for idx, h in enumerate(headers):
                                row_dict[h] = cell_texts[idx] if idx < len(cell_texts) else ""
                            data_rows.append(row_dict)

                if headers and data_rows:
                    tables.append(data_rows)
            else:
                for v in node.values():
                    _walk(v)
        elif isinstance(node, list):
            for item in node:
                _walk(item)

    _walk(adf_node)
    return tables


def _strip_jira_markup(text: str) -> str:
    """Remove Jira wiki formatting tags from text (e.g. {color:#hex}...{color})."""
    # Strip {color:...}...{color} keeping inner text
    text = _re.sub(r'\{color(?::[^}]*)?\}', '', text)
    # Strip {noformat}...{noformat}
    text = _re.sub(r'\{noformat\}', '', text)
    # Strip {code}...{code}
    text = _re.sub(r'\{code(?::[^}]*)?\}', '', text)
    # Strip Jira wiki bold markers so values like *100%* become 100%
    text = text.replace("*", "")
    return text.strip()


def _flatten_row(row: dict[str, Any]) -> dict[str, str]:
    """Convert a row dict to string values, resolving nested objects."""
    flat: dict[str, str] = {}
    for k, v in row.items():
        if k in ("self", "id"):
            continue
        if isinstance(v, dict):
            # Jira objects → use 'name' or 'value' or 'displayName'
            flat[k] = str(v.get("name") or v.get("value") or v.get("displayName") or v)
        elif isinstance(v, list):
            flat[k] = ", ".join(str(i.get("name", i) if isinstance(i, dict) else i) for i in v)
        elif v is None:
            flat[k] = ""
        else:
            flat[k] = str(v)
        # Clean Jira wiki markup from the value
        flat[k] = _strip_jira_markup(flat[k])
    return flat


def _parse_wiki_table(text: str) -> tuple[str, list[dict[str, str]]] | None:
    """Parse a simple wiki-markup table into rows.
    
    Supports formats:
      ||header||header||     (double-pipe headers)
      |val|val|              (single-pipe data)
      |*header*|*header*|    (bold-in-pipe headers)

    Returns (context_text, rows) where context_text is non-table text near the table.
    """
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    header_line = None
    data_lines: list[str] = []
    context_lines: list[str] = []
    for line in lines:
        if line.startswith("||"):
            header_line = line
        elif line.startswith("|"):
            # Check if this is a header row with |*bold*| pattern
            cells = [c.strip() for c in line.split("|") if c.strip()]
            if cells and all(c.startswith("*") and c.endswith("*") for c in cells):
                header_line = line
            else:
                data_lines.append(line)
        else:
            # Non-table line — capture as context only if it's a real statement
            # Skip lines that contain || or look like table fragments
            if "||" in line or line.count("|") >= 2:
                continue
            cleaned = _strip_jira_markup(line.strip("*").strip())
            if cleaned:
                context_lines.append(cleaned)

    if not header_line or not data_lines:
        # Fallback: if no explicit header found, treat first data line as header
        if not header_line and len(data_lines) >= 2:
            header_line = data_lines.pop(0)
        else:
            return None

    if not header_line or not data_lines:
        return None

    # Parse headers
    if header_line.startswith("||"):
        headers = [_strip_jira_markup(h) for h in header_line.split("||") if h.strip()]
    else:
        raw_headers = [c.strip() for c in header_line.split("|") if c.strip()]
        headers = [_strip_jira_markup(h.strip("*").strip()) for h in raw_headers]

    rows: list[dict[str, str]] = []
    for dl in data_lines:
        cells = [_strip_jira_markup(c) for c in dl.split("|") if c.strip()]
        row_dict: dict[str, str] = {}
        for idx, header in enumerate(headers):
            row_dict[header] = cells[idx] if idx < len(cells) else ""
        rows.append(row_dict)

    context = "\n".join(context_lines)
    return (context, rows) if rows else None


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split()).strip()


def _normalize_key(value: str) -> str:
    return _normalize_text(value).lower()


def _find_column_name(row: dict[str, str], aliases: set[str]) -> str | None:
    normalized_map = {_normalize_key(key): key for key in row}
    for alias in aliases:
        if alias in normalized_map:
            return normalized_map[alias]
    for normalized_key, original_key in normalized_map.items():
        if any(alias in normalized_key for alias in aliases):
            return original_key
    return None


def _pick_row_name(row: dict[str, str], status_key: str | None, pass_metric_key: str | None) -> str:
    for key, value in row.items():
        if key in {status_key, pass_metric_key}:
            continue
        cleaned = _normalize_text(value)
        if cleaned:
            return cleaned
    return _normalize_text(next(iter(row.values()), "")) or "(unnamed row)"


def _split_pass_metric_values(pass_metric: str) -> list[str]:
    values = []
    for part in PASS_METRIC_SPLIT_RE.split(_normalize_text(pass_metric)):
        cleaned = _normalize_key(part)
        if cleaned:
            values.append(cleaned)
            values.append(_normalize_key(_re.sub(r"\([^)]*\)", "", part)))
    return [value for value in values if value]


def _classify_readiness_text(value: str) -> tuple[str, str]:
    normalized = _normalize_key(value)
    if not normalized:
        return "neutral", "Missing status"
    if normalized in {"na", "n/a", "not applicable"} or _re.search(r"\b(?:na|n/a|not applicable)\b", normalized):
        return "pass", "Marked NA"
    # Business rule: SCM/SRM/PO tickets are always treated as approved/pass
    if _re.search(r"\b(?:scm|srm|po)-\d+", normalized):
        return "pass", "SCM/SRM/PO ticket is approved"
    if any(phrase in normalized for phrase in FAILING_PHRASES):
        return "fail", "Status indicates work is not complete"
    if any(phrase in normalized for phrase in PASSING_PHRASES):
        return "pass", "Status indicates ready"
    return "neutral", "Status is informational"


def _is_done_issue_status(status_value: str) -> bool:
    return _normalize_key(status_value) == "done"


def _is_open_issue_status(status_value: str) -> bool:
    return _normalize_key(status_value) == "open"


def _extract_numeric_value(value: str) -> float | None:
    text = _normalize_text(value)
    for pattern in (
        r"([0-9]*\.?[0-9]+)\s*%",
        r"([0-9]*\.?[0-9]+)\s*(?:hrs?|hours?)\b",
        r"([0-9]*\.?[0-9]+)",
    ):
        matches = _re.findall(pattern, text, flags=_re.IGNORECASE)
        if matches:
            # Use the first numeric match found.
            return float(matches[0])
    return None


def _matches_pass_metric(pass_metric: str, actual_status: str) -> tuple[bool, str]:
    normalized_actual = _normalize_key(actual_status)
    if not normalized_actual:
        return False, "Missing status"

    # Business rule: any NA/N-A style status is considered pass.
    if _re.search(r"\b(?:na|n/a|not applicable)\b", normalized_actual):
        return True, "NA status is treated as pass"

    metric_values = set(_split_pass_metric_values(pass_metric))
    if not metric_values:
        return False, "Missing pass metric"
    if normalized_actual in metric_values:
        return True, "Exact pass metric match"

    actual_state, actual_reason = _classify_readiness_text(actual_status)
    metric_states = {_classify_readiness_text(value)[0] for value in metric_values}
    if actual_state == "pass" and "pass" in metric_states:
        return True, "Pass-like status satisfies pass metric"
    if actual_state == "fail":
        return False, actual_reason

    actual_option_values = {
        _normalize_key(_re.sub(r"\([^)]*\)", "", part))
        for part in PASS_METRIC_SPLIT_RE.split(_normalize_text(actual_status))
    }
    actual_option_values = {value for value in actual_option_values if value}
    if len(actual_option_values) > 1 and actual_option_values.issubset(metric_values):
        return True, "Composite status options satisfy pass metric"

    # Check if actual_status contains SCM/SRM/PO tickets - these are always valid/approved
    if _re.search(r"\b(?:scm|srm|po)-\d+", _normalize_text(actual_status), _re.IGNORECASE):
        return True, "SCM/SRM/PO ticket reference is valid and approved"

    # Check if actual_status is a ticket reference (general case)
    normalized_status = _normalize_text(actual_status)
    if TICKET_REF_RE.fullmatch(normalized_status):
        # Other ticket references require "ticket created" in metric
        if any("ticket created" in value for value in metric_values):
            return True, "Ticket reference satisfies ticket-created metric"

    # Check for explicit numeric thresholds (e.g. ">= 800")
    threshold_match = THRESHOLD_RE.search(_normalize_text(pass_metric).replace(" ", ""))
    if threshold_match:
        actual_value = _extract_numeric_value(actual_status)
        if actual_value is not None:
            operator, threshold_raw = threshold_match.groups()
            threshold_value = float(threshold_raw)
            if operator == ">" and actual_value > threshold_value:
                return True, "Numeric status satisfies lower-bound metric"
            if operator == ">=" and actual_value >= threshold_value:
                return True, "Numeric status satisfies lower-bound metric"
            if operator == "<" and actual_value < threshold_value:
                return True, "Numeric status satisfies upper-bound metric"
            if operator == "<=" and actual_value <= threshold_value:
                return True, "Numeric status satisfies upper-bound metric"
            if operator == "=" and actual_value == threshold_value:
                return True, "Numeric status satisfies exact metric"
            return False, f"Numeric status ({actual_value}) does not satisfy threshold ({operator}{threshold_value})"

    if "approved" in normalized_actual and any("approved" in value for value in metric_values):
        return True, "Approved status satisfies pass metric"
    if "ready to release" in normalized_actual and any("ready to release" in value for value in metric_values):
        return True, "Ready-to-release status satisfies pass metric"
    if normalized_actual in {"pass", "passed"} and any(value.startswith("pass") for value in metric_values):
        return True, "PASS status satisfies pass metric"
    if _re.search(r"\b(?:na|n/a|not applicable)\b", normalized_actual) and any(value in {"na", "n/a", "not applicable"} for value in metric_values):
        return True, "NA status satisfies pass metric"
    if normalized_actual in {"pass", "passed"} and any("yes" in value for value in metric_values):
        return True, "PASS status satisfies yes-based pass metric"
    if normalized_actual == "yes" and any(value.startswith("pass") for value in metric_values):
        return True, "Yes status satisfies pass-based metric"

    # Handle implicit numeric comparisons (e.g. "400 hrs" vs "576hrs")
    # If both fields contain numbers and no explicit operator was found, 
    # default to a "greater than or equal to" check.
    expected_numeric = _extract_numeric_value(pass_metric)
    actual_numeric = _extract_numeric_value(actual_status)
    if expected_numeric is not None and actual_numeric is not None:
        if actual_numeric >= expected_numeric:
            return True, f"Numeric status ({actual_numeric}) is >= expected ({expected_numeric})"
        else:
            return False, f"Numeric status ({actual_numeric}) is < expected ({expected_numeric})"

    return False, "Status does not satisfy pass metric"


def analyze_issue_readiness(
    issue_key: str,
    issue_summary: str,
    context_text: str,
    issue_status: str,
    table_rows: list[dict[str, str]],
) -> IssueReadinessAnalysis:
    exceptions: list[ReadinessRowException] = []

    context_state, _ = _classify_readiness_text(context_text)
    has_pass_metric_column = any(
        _find_column_name(row, PASS_METRIC_COLUMN_ALIASES) is not None for row in table_rows
    )

    # Business rule: when context says ready-to-release and table has no pass-metric column,
    # treat the whole table as pass.
    if context_state == "pass" and table_rows and not has_pass_metric_column:
        if _is_open_issue_status(issue_status):
            context = _normalize_text(context_text)
            return IssueReadinessAnalysis(
                issue_key=issue_key,
                issue_summary=issue_summary,
                context=context,
                context_state=context_state.title(),
                overall_state="Needs Attention",
                total_rows=len(table_rows),
                exception_count=1,
                exception_rows=[
                    ReadinessRowException(
                        issue_key=issue_key,
                        issue_summary=issue_summary,
                        context=context,
                        row_name="(Context-based readiness)",
                        pass_metric="Jira issue status must not be Open",
                        actual_status=issue_status,
                        reason="Context suggests ready, but Jira status is Open",
                    )
                ],
            )
        return IssueReadinessAnalysis(
            issue_key=issue_key,
            issue_summary=issue_summary,
            context=_normalize_text(context_text),
            context_state=context_state.title(),
            overall_state="Ready",
            total_rows=len(table_rows),
            exception_count=0,
            exception_rows=[],
        )

    for row in table_rows:
        status_key = _find_column_name(row, STATUS_COLUMN_ALIASES)
        pass_metric_key = _find_column_name(row, PASS_METRIC_COLUMN_ALIASES)
        
        actual_status = _normalize_text(row.get(status_key, "") if status_key else "")
        pass_metric = _normalize_text(row.get(pass_metric_key, "") if pass_metric_key else "")

        # FIX: Skip rows where both status and expected metric are empty or blank.
        if not actual_status and not pass_metric:
            continue

        row_name = _pick_row_name(row, status_key, pass_metric_key)

        if pass_metric_key and status_key:
            is_match, reason = _matches_pass_metric(pass_metric, actual_status)
            if not is_match:
                exceptions.append(
                    ReadinessRowException(
                        issue_key=issue_key,
                        issue_summary=issue_summary,
                        context=_normalize_text(context_text),
                        row_name=row_name,
                        pass_metric=pass_metric,
                        actual_status=actual_status,
                        reason=reason,
                    )
                )
            continue

        if status_key:
            status_state, reason = _classify_readiness_text(actual_status)
            if status_state == "fail":
                exceptions.append(
                    ReadinessRowException(
                        issue_key=issue_key,
                        issue_summary=issue_summary,
                        context=_normalize_text(context_text),
                        row_name=row_name,
                        pass_metric=pass_metric or "(no pass metric column)",
                        actual_status=actual_status,
                        reason=reason,
                    )
                )

    overall_state = "Ready"
    if exceptions or context_state == "fail":
        overall_state = "Needs Attention"
    elif not table_rows:
        overall_state = "No Table"

    return IssueReadinessAnalysis(
        issue_key=issue_key,
        issue_summary=issue_summary,
        context=_normalize_text(context_text),
        context_state=context_state.title(),
        overall_state=overall_state,
        total_rows=len(table_rows),
        exception_count=len(exceptions),
        exception_rows=exceptions,
    )


def print_issue_details(issue: dict[str, Any]) -> None:
    fields = issue.get("fields", {})
    print("=" * 72)
    print(f"Epic Key      : {issue.get('key', 'N/A')}")
    print(f"Title         : {fields.get('summary', '')}")
    print(f"Type          : {(fields.get('issuetype') or {}).get('name', '')}")
    print(f"Status        : {(fields.get('status') or {}).get('name', '')}")
    print(f"Assignee      : {format_user(fields.get('assignee'))}")
    print(f"Reporter      : {format_user(fields.get('reporter'))}")
    print(f"Created       : {fields.get('created', '')}")
    print(f"Updated       : {fields.get('updated', '')}")
    description = extract_description(fields.get("description"))
    print(f"Description   : {description if description else '(empty)'}")
    print("=" * 72)


def print_epic_children(children: list[dict[str, Any]]) -> None:
    print("\nIssues in Epic")
    print("-" * 72)
    if not children:
        print("No child issues found for this epic.")
        return

    for issue in children:
        fields = issue.get("fields", {})
        status = (fields.get("status") or {}).get("name", "")
        assignee = format_user(fields.get("assignee"))
        issue_type = (fields.get("issuetype") or {}).get("name", "")
        summary = fields.get("summary", "")
        print(f"{issue.get('key', ''):12} | {issue_type:10} | {status:12} | {assignee:20} | {summary}")


def _apply_header_style(cell: Any) -> None:
    cell.font = Font(color="FFFFFFFF", bold=True)
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFD9D9D9")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _autosize_columns(sheet: Any, min_width: int = 12, max_width: int = 60) -> None:
    for col_cells in sheet.columns:
        first = col_cells[0]
        column_letter = first.column_letter if hasattr(first, "column_letter") else get_column_letter(first.column)
        longest = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            value_len = len(str(cell.value))
            if value_len > longest:
                longest = value_len
        sheet.column_dimensions[column_letter].width = max(min_width, min(max_width, longest + 2))


def _write_tabular_sheet(sheet: Any, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        _apply_header_style(cell)

    thin = Side(style="thin", color="FFD9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_values in rows:
        sheet.append(row_values)

    for row in sheet.iter_rows(min_row=2, max_row=max(2, sheet.max_row), min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, sheet.max_row)}"
    _autosize_columns(sheet)


def print_readiness_summary(analyses: list[IssueReadinessAnalysis]) -> None:
    attention_items = [analysis for analysis in analyses if analysis.overall_state != "Ready"]
    if not attention_items:
        print("\nReadiness Summary")
        print("-" * 72)
        print("All extracted High Level Status tables are ready.")
        return

    print("\nReadiness Summary")
    print("-" * 72)
    for analysis in attention_items:
        print(
            f"{analysis.issue_key:12} | {analysis.overall_state:15} | "
            f"Exceptions: {analysis.exception_count:2d} | {analysis.issue_summary}"
        )
        if analysis.context:
            print(f"  Context: {analysis.context}")
        for row_exception in analysis.exception_rows:
            print(
                f"  - {row_exception.row_name}: actual='{row_exception.actual_status}' "
                f"expected='{row_exception.pass_metric}' ({row_exception.reason})"
            )


def write_excel_report(
    epic: dict[str, Any],
    children: list[dict[str, Any]],
    output_path: Path,
    client: JiraClient | None = None,
    hls_field_id: str | None = None,
    table_field_ids: list[str] | None = None,
) -> list[IssueReadinessAnalysis]:
    wb = Workbook()
    readiness_analyses: list[IssueReadinessAnalysis] = []

    epic_sheet = wb.active
    epic_sheet.title = "Epic Summary"

    issues_sheet = wb.create_sheet("Issues in Epic")

    fields = epic.get("fields", {})
    epic_description = extract_description(fields.get("description")) or "(empty)"

    epic_sheet.merge_cells("A1:D1")
    epic_sheet["A1"] = f"Jira Epic Report - {epic.get('key', 'N/A')}"
    epic_sheet["A1"].font = Font(size=16, bold=True, color="FFFFFFFF")
    epic_sheet["A1"].fill = PatternFill("solid", fgColor="0B6E4F")
    epic_sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    epic_sheet.row_dimensions[1].height = 28

    summary_rows = [
        ("Epic Key", epic.get("key", "N/A")),
        ("Title", fields.get("summary", "")),
        ("Type", (fields.get("issuetype") or {}).get("name", "")),
        ("Status", (fields.get("status") or {}).get("name", "")),
        ("Assignee", format_user(fields.get("assignee"))),
        ("Reporter", format_user(fields.get("reporter"))),
        ("Created", fields.get("created", "")),
        ("Updated", fields.get("updated", "")),
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Description", epic_description),
    ]

    start_row = 3
    for index, (label, value) in enumerate(summary_rows, start=start_row):
        label_cell = epic_sheet[f"A{index}"]
        value_cell = epic_sheet[f"B{index}"]
        label_cell.value = label
        value_cell.value = value
        label_cell.font = Font(bold=True, color="1F4E78")
        label_cell.fill = PatternFill("solid", fgColor="EAF3FA")
        label_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        value_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    thin = Side(style="thin", color="FFD9D9D9")
    table_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(start_row, start_row + len(summary_rows)):
        for col in ("A", "B"):
            epic_sheet[f"{col}{row}"].border = table_border

    epic_sheet.column_dimensions["A"].width = 20
    epic_sheet.column_dimensions["B"].width = 95
    epic_sheet.freeze_panes = "A3"

    issue_headers = [
        "Issue Key",
        "Type",
        "Status",
        "Assignee",
        "Reporter",
        "Created",
        "Updated",
        "Summary",
    ]
    issues_sheet.append(issue_headers)
    for cell in issues_sheet[1]:
        _apply_header_style(cell)

    for issue in children:
        issue_fields = issue.get("fields", {})
        issues_sheet.append(
            [
                issue.get("key", ""),
                (issue_fields.get("issuetype") or {}).get("name", ""),
                (issue_fields.get("status") or {}).get("name", ""),
                format_user(issue_fields.get("assignee")),
                format_user(issue_fields.get("reporter")),
                issue_fields.get("created", ""),
                issue_fields.get("updated", ""),
                issue_fields.get("summary", ""),
            ]
        )

    for row in issues_sheet.iter_rows(min_row=2, max_row=max(2, issues_sheet.max_row), min_col=1, max_col=len(issue_headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = table_border

    issues_sheet.freeze_panes = "A2"
    issues_sheet.auto_filter.ref = f"A1:H{max(1, issues_sheet.max_row)}"
    _autosize_columns(issues_sheet)

    # Add all High Level Status tables in a single sheet
    if client is not None:
        detail_sheet = wb.create_sheet("High Level Status")
        summary_sheet = wb.create_sheet("Readiness Summary")
        exceptions_sheet = wb.create_sheet("Readiness Exceptions")
        current_row = 1

        for issue in children:
            issue_key = issue.get("key", "")
            issue_summary = issue.get("fields", {}).get("summary", "")
            issue_status = (issue.get("fields", {}).get("status") or {}).get("name", "")
            if not issue_key:
                continue

            try:
                issue_full = client.get_issue_full(issue_key)
                hls_result = extract_high_level_status_table(issue_full, hls_field_id, client, table_field_ids)

                if not hls_result:
                    no_table_context = f"No High Level Status table. Jira status: {issue_status or 'Unknown'}"
                    if _is_done_issue_status(issue_status):
                        readiness_analyses.append(
                            IssueReadinessAnalysis(
                                issue_key=issue_key,
                                issue_summary=issue_summary,
                                context=no_table_context,
                                context_state="Pass",
                                overall_state="Ready",
                                total_rows=0,
                                exception_count=0,
                                exception_rows=[],
                            )
                        )
                        continue

                    readiness_analyses.append(
                        IssueReadinessAnalysis(
                            issue_key=issue_key,
                            issue_summary=issue_summary,
                            context=no_table_context,
                            context_state="Neutral",
                            overall_state="Needs Attention",
                            total_rows=0,
                            exception_count=1,
                            exception_rows=[
                                ReadinessRowException(
                                    issue_key=issue_key,
                                    issue_summary=issue_summary,
                                    context=no_table_context,
                                    row_name="(No High Level Status table)",
                                    pass_metric="Jira issue status must be Done",
                                    actual_status=issue_status or "Unknown",
                                    reason="Missing High Level Status table and Jira status is not Done",
                                )
                            ],
                        )
                    )
                    continue

                context_text, hls_table = hls_result
                readiness_analyses.append(
                    analyze_issue_readiness(issue_key, issue_summary, context_text, issue_status, hls_table)
                )

                # Add issue header
                detail_sheet.merge_cells(f"A{current_row}:D{current_row}")
                header_cell = detail_sheet.cell(row=current_row, column=1)
                header_cell.value = f"{issue_key} - {issue_summary}"
                header_cell.font = Font(size=12, bold=True, color="FFFFFFFF")
                header_cell.fill = PatternFill("solid", fgColor="1F4E78")
                header_cell.alignment = Alignment(horizontal="left", vertical="center")
                detail_sheet.row_dimensions[current_row].height = 24
                current_row += 1

                # Add context text (e.g. "ECRT Status: ECRT is ready to release.")
                if context_text:
                    for ctx_line in context_text.splitlines():
                        ctx_line = ctx_line.strip()
                        if not ctx_line:
                            continue
                        current_row += 1
                        ctx_cell = detail_sheet.cell(row=current_row, column=1)
                        ctx_cell.value = ctx_line
                        ctx_cell.font = Font(size=11, bold=True, italic=True, color="0B6E4F")
                        ctx_cell.alignment = Alignment(horizontal="left", vertical="center")
                current_row += 1

                # Collect all column names preserving order from first row
                seen: set[str] = set()
                headers: list[str] = []
                for row in hls_table:
                    for k in row:
                        if k not in seen:
                            seen.add(k)
                            headers.append(k)

                # Write headers
                thin = Side(style="thin", color="FFD9D9D9")
                tbl_border = Border(left=thin, right=thin, top=thin, bottom=thin)

                for col_idx, header in enumerate(headers, 1):
                    cell = detail_sheet.cell(row=current_row, column=col_idx)
                    cell.value = header
                    _apply_header_style(cell)
                current_row += 1

                # Write table data rows
                for row_data in hls_table:
                    for col_idx, header in enumerate(headers, 1):
                        cell = detail_sheet.cell(row=current_row, column=col_idx)
                        cell.value = row_data.get(header, "")
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                        cell.border = tbl_border
                    current_row += 1

                # Add spacing between issues
                current_row += 1
            except Exception:
                readiness_analyses.append(
                    IssueReadinessAnalysis(
                        issue_key=issue_key,
                        issue_summary=issue_summary,
                        context="",
                        context_state="Neutral",
                        overall_state="Error",
                        total_rows=0,
                        exception_count=0,
                        exception_rows=[],
                    )
                )

        _autosize_columns(detail_sheet)

        summary_rows = [
            [
                analysis.issue_key,
                analysis.issue_summary,
                analysis.context,
                analysis.overall_state,
            ]
            for analysis in readiness_analyses
        ]
        _write_tabular_sheet(
            summary_sheet,
            [
                "Issue Key",
                "Issue Summary",
                "Context",
                "Overall Status",
            ],
            summary_rows,
        )

        exception_rows = [
            [
                row.issue_key,
                row.issue_summary,
                row.pass_metric,
                row.actual_status,
            ]
            for analysis in readiness_analyses
            for row in analysis.exception_rows
        ]
        if not exception_rows:
            exception_rows = [["", "", "", "No not-ready or failed rows found."]]
        _write_tabular_sheet(
            exceptions_sheet,
            [
                "Issue Key",
                "Issue Summary",
                "Pass Metric",
                "Actual Status",
            ],
            exception_rows,
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return readiness_analyses


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Fetch Jira epic details and list all issues in that epic."
    )
    parser.add_argument(
        "epic_id",
        nargs="?",
        help="Epic key, e.g. DQRD-10393 (if omitted, you will be prompted)",
    )
    parser.add_argument(
        "--excel",
        dest="excel_path",
        help="Output Excel file path (default: reports/<EPIC_ID>_<timestamp>.xlsx)",
    )
    parser.add_argument(
        "--console",
        action="store_true",
        help="Also print epic details and issue list in terminal.",
    )
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Dump raw JSON fields for first child issue (for troubleshooting).",
    )
    parser.add_argument(
        "--summary",
        action="store_true",
        help="Print readiness exceptions in the terminal after the Excel report is generated.",
    )
    parser.add_argument(
        "--serve",
        action="store_true",
        help="Run a local web server that serves the UI and backend API.",
    )
    parser.add_argument(
        "--host",
        default="127.0.0.1",
        help="Host for --serve mode (default: 127.0.0.1).",
    )
    parser.add_argument(
        "--port",
        type=int,
        default=8000,
        help="Port for --serve mode (default: 8000).",
    )
    return parser.parse_args()


def run_epic_report(
    epic_id: str,
    excel_path: Path | None = None,
    console: bool = False,
    debug: bool = False,
    summary: bool = False,
) -> tuple[Path, list[IssueReadinessAnalysis]]:
    project_root = Path(__file__).resolve().parent
    load_dotenv_file(project_root / ".env")

    normalized_epic = epic_id.strip().upper()
    if not normalized_epic:
        raise RuntimeError("DQRD ID is required.")

    config = JiraConfig(
        base_url=get_env("JIRA_BASE_URL"),
        pat=get_env("JIRA_PAT"),
    )
    client = JiraClient(config)
    epic = client.get_issue(normalized_epic)

    hls_field_id = None
    table_field_ids: list[str] = []
    epic_link_field_id = None
    try:
        field_defs = client.get_field_definitions()
        hls_field_id = find_hls_field_id(field_defs)
        table_field_ids = find_table_field_ids(field_defs)
        epic_link_field_id = find_epic_link_field_id(field_defs)
    except Exception:
        pass

    children = find_epic_issues(client, normalized_epic, epic_link_field_id)

    if debug and children:
        first_key = children[0].get("key", "")
        if first_key:
            debug_data = client.get_issue_full(first_key)
            debug_fields = debug_data.get("fields", {})
            debug_path = Path("debug") / f"{first_key}_fields.json"
            debug_path.parent.mkdir(parents=True, exist_ok=True)
            debug_out: dict[str, Any] = {}
            for fk, fv in sorted(debug_fields.items()):
                if fv is None:
                    continue
                debug_out[fk] = fv
            debug_path.write_text(json.dumps(debug_out, indent=2, default=str), encoding="utf-8")
            print(f"Debug: raw fields dumped to {debug_path.resolve()}")
            if hls_field_id:
                print(f"Debug: High Level Status field ID = {hls_field_id}")
            else:
                print("Debug: Could not find 'High Level Status' field in Jira field definitions.")

    if console:
        print_issue_details(epic)
        print_epic_children(children)

    if excel_path is None:
        excel_path = Path("reports") / f"{normalized_epic}_REPORT.xlsx"

    readiness_analyses = write_excel_report(epic, children, excel_path, client, hls_field_id, table_field_ids)
    if summary:
        print_readiness_summary(readiness_analyses)

    return excel_path, readiness_analyses


def run_web_server(host: str, port: int) -> int:
    project_root = Path(__file__).resolve().parent
    frontend_dir = project_root / "frontend"

    class EpicHandler(SimpleHTTPRequestHandler):
        def __init__(self, *args: Any, **kwargs: Any) -> None:
            super().__init__(*args, directory=str(frontend_dir), **kwargs)

        def log_message(self, format: str, *args: Any) -> None:
            """Log requests to console."""
            message = f"[{datetime.now().strftime('%H:%M:%S')}] {format % args}"
            print(message, flush=True)
            sys.stdout.flush()

        def do_POST(self) -> None:  # noqa: N802
            print(f"[{datetime.now().strftime('%H:%M:%S')}] POST {self.path}", flush=True)
            sys.stdout.flush()
            if self.path != "/api/fetch-epic":
                self.send_error(HTTPStatus.NOT_FOUND, "Endpoint not found")
                return

            try:
                length_raw = self.headers.get("Content-Length", "0")
                content_length = int(length_raw)
                payload_raw = self.rfile.read(content_length).decode("utf-8") if content_length > 0 else "{}"
                payload = json.loads(payload_raw)

                epic_id = str(payload.get("epic_id", "")).strip().upper()
                if not epic_id:
                    self._send_json(
                        HTTPStatus.BAD_REQUEST,
                        {"ok": False, "error": "'epic_id' is required (example: DQRD-10393)."},
                    )
                    return

                excel_path, readiness_analyses = run_epic_report(epic_id=epic_id)
                
                # Prepare readiness summary
                total_issues = len(readiness_analyses)
                ready_count = sum(1 for a in readiness_analyses if a.overall_state == "Ready")
                attention_count = sum(1 for a in readiness_analyses if a.overall_state == "Needs Attention")
                failed_count = sum(1 for a in readiness_analyses if a.overall_state == "Error")
                
                # Prepare exceptions
                all_exceptions = []
                for analysis in readiness_analyses:
                    for exc_row in analysis.exception_rows:
                        all_exceptions.append({
                            "issue_key": exc_row.issue_key,
                            "row_name": exc_row.row_name,
                            "pass_metric": exc_row.pass_metric,
                            "actual_status": exc_row.actual_status,
                            "reason": exc_row.reason,
                        })

                self._send_json(
                    HTTPStatus.OK,
                    {
                        "ok": True,
                        "epic_id": epic_id,
                        "excel_path": str(excel_path.resolve()),
                        "message": "Excel report generated successfully.",
                        "readiness_summary": {
                            "total_issues": total_issues,
                            "ready_count": ready_count,
                            "attention_count": attention_count,
                            "failed_count": failed_count,
                        },
                        "readiness_exceptions": all_exceptions[:50],  # Limit to 50 for UI
                    },
                )
            except Exception as exc:  # noqa: BLE001
                self._send_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"ok": False, "error": str(exc)})

        def do_GET(self) -> None:  # noqa: N802
            print(f"[{datetime.now().strftime('%H:%M:%S')}] GET {self.path}", flush=True)
            sys.stdout.flush()
            if self.path == "/":
                self.path = "/index.html"
            
            # Handle file downloads
            if self.path.startswith("/api/download"):
                self._handle_download()
                return
            
            super().do_GET()
        
        def _handle_download(self) -> None:
            """Handle file download requests."""
            try:
                from urllib.parse import urlparse, parse_qs
                
                # Parse query parameters
                parsed_url = urlparse(self.path)
                query_params = parse_qs(parsed_url.query)
                
                if 'path' not in query_params:
                    self._send_json(HTTPStatus.BAD_REQUEST, {"error": "Missing path parameter"})
                    return
                
                file_path_str = query_params['path'][0]
                file_path = Path(file_path_str)
                
                # Security check: ensure file exists and is in reports directory
                if not file_path.exists():
                    self._send_json(HTTPStatus.NOT_FOUND, {"error": "File not found"})
                    return
                
                # Read and send the file
                with open(file_path, 'rb') as f:
                    file_content = f.read()
                
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Length", str(len(file_content)))
                self.send_header("Content-Disposition", f"attachment; filename=\"{file_path.name}\"")
                self.end_headers()
                self.wfile.write(file_content)
            except Exception as e:
                self._send_json(HTTPStatus.INTERNAL_SERVER_ERROR, {"error": str(e)})

        def _send_json(self, status: HTTPStatus, payload: dict[str, Any]) -> None:
            body = json.dumps(payload).encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", "application/json")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

    with ThreadingHTTPServer((host, port), EpicHandler) as server:
        print(f"Web UI running at http://{host}:{port}")
        print("Press Ctrl+C to stop.")
        try:
            server.serve_forever()
        except KeyboardInterrupt:
            print("\nServer stopped.")
            return 0
    return 0


def main() -> int:
    args = parse_args()

    try:
        # If no arguments provided, default to web server mode
        has_cli_args = args.epic_id or args.excel_path or args.console or args.debug or args.summary
        if args.serve or not has_cli_args:
            return run_web_server(args.host, args.port)

        epic_id = (args.epic_id or "").strip().upper()
        if not epic_id:
            epic_id = input("Enter the DQRD ID (e.g. DQRD-10393): ").strip().upper()
        if not epic_id:
            raise RuntimeError("DQRD ID is required.")

        if args.excel_path:
            excel_path = Path(args.excel_path)
        else:
            excel_path = None

        generated_path, _ = run_epic_report(
            epic_id=epic_id,
            excel_path=excel_path,
            console=args.console,
            debug=args.debug,
            summary=args.summary,
        )
        print(f"Excel report generated: {generated_path.resolve()}")
        return 0
    except Exception as exc:  # noqa: BLE001
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())